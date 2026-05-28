from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.models import Lead
from utils.text import safe_filename_part


def leads_to_dataframe(leads: list[Lead]) -> pd.DataFrame:
    rows = [lead.export_dict() for lead in leads]
    columns = [
        "Business Name", "Category", "Address", "Website Status", "Website URL",
        "Contact Number", "WhatsApp", "Email", "Instagram", "Facebook",
        "Google Maps Link", "Rating", "Review Count",
    ]
    return pd.DataFrame(rows, columns=columns).fillna("N/A")


def build_excel(leads: list[Lead]) -> bytes:
    df = leads_to_dataframe(leads)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        sheet = writer.book["Leads"]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_len + 2, 14), 55)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    return output.getvalue()


def export_filename(location: str, category: str) -> str:
    return f"{safe_filename_part(location)}_{safe_filename_part(category)}_leads.xlsx"
